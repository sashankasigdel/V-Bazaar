from rest_framework import generics, permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django.contrib.auth import get_user_model
from django.db.models import Count, Sum, Q
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from apps.businesses.models import Business, BusinessCategory, Advertisement
from apps.businesses.qr import generate_qr_card, business_target_url
from apps.businesses.export import businesses_to_excel
from apps.orders.models import Order
from apps.reviews.models import Review
from apps.products.models import Product
import io
import zipfile

User = get_user_model()


class IsAdminUser(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and (request.user.is_staff or request.user.role == 'admin')


@api_view(['GET'])
@permission_classes([IsAdminUser])
def admin_stats(request):
    now = timezone.now()
    month_ago = now - timedelta(days=30)
    week_ago = now - timedelta(days=7)

    total_users = User.objects.count()
    total_businesses = Business.objects.count()
    active_businesses = Business.objects.filter(status='active').count()
    pending_businesses = Business.objects.filter(status='pending').count()
    total_orders = Order.objects.count()
    total_revenue = Order.objects.filter(status='delivered').aggregate(Sum('total'))['total__sum'] or 0
    total_reviews = Review.objects.count()
    total_customers = User.objects.filter(role='customer').count()
    total_owners = User.objects.filter(role='business_owner').count()

    orders_this_month = Order.objects.filter(created_at__gte=month_ago).count()
    orders_this_week = Order.objects.filter(created_at__gte=week_ago).count()
    revenue_this_month = Order.objects.filter(status='delivered', created_at__gte=month_ago).aggregate(Sum('total'))['total__sum'] or 0
    new_users_this_month = User.objects.filter(date_joined__gte=month_ago).count()
    new_businesses_this_month = Business.objects.filter(created_at__gte=month_ago).count()

    return Response({
        'total_users': total_users,
        'total_customers': total_customers,
        'total_owners': total_owners,
        'total_businesses': total_businesses,
        'active_businesses': active_businesses,
        'pending_businesses': pending_businesses,
        'total_orders': total_orders,
        'total_revenue': str(total_revenue),
        'total_reviews': total_reviews,
        'orders_this_month': orders_this_month,
        'orders_this_week': orders_this_week,
        'revenue_this_month': str(revenue_this_month),
        'new_users_this_month': new_users_this_month,
        'new_businesses_this_month': new_businesses_this_month,
    })


@api_view(['GET'])
@permission_classes([IsAdminUser])
def admin_businesses(request):
    from apps.businesses.serializers import BusinessListSerializer
    status_filter = request.query_params.get('status', '')
    search = request.query_params.get('search', '')
    qs = Business.objects.all().select_related('category', 'owner')
    if status_filter:
        qs = qs.filter(status=status_filter)
    if search:
        q = Q(name__icontains=search)
        if search.isdigit():
            q |= Q(id=int(search))
        qs = qs.filter(q)
    qs = qs.order_by('-created_at')

    data = []
    for b in qs:
        data.append({
            'id': b.id,
            'name': b.name,
            'slug': b.slug,
            'owner_id': b.owner_id,
            'owner_email': b.owner.email,
            'owner_name': b.owner.get_full_name(),
            'owner_phone': b.owner.phone,
            'category': b.category.name if b.category else '',
            'category_id': b.category_id,
            'status': b.status,
            'is_featured': b.is_featured,
            'is_verified': b.is_verified,
            'accepts_orders': b.accepts_orders,
            'average_rating': str(b.average_rating),
            'total_orders': b.total_orders,
            'total_views': b.total_views,
            'city': b.city,
            'phone': b.phone,
            'has_banner': bool(b.banner),
            'parent_id': b.parent_id,
            'parent_name': b.parent.name if b.parent else '',
            'business_code': b.business_code,
            'branch_code': b.branch_code,
            'created_at': b.created_at.isoformat(),
        })
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def admin_business_detail(request, pk):
    from apps.businesses.serializers import BusinessDetailSerializer
    try:
        business = Business.objects.get(pk=pk)
    except Business.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)
    data = BusinessDetailSerializer(business, context={'request': request}).data
    data['owner_email'] = business.owner.email
    return Response(data)


@api_view(['PATCH'])
@permission_classes([IsAdminUser])
def admin_update_business(request, pk):
    from apps.businesses.serializers import AdminBusinessCreateSerializer
    try:
        business = Business.objects.get(pk=pk)
    except Business.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)

    serializer = AdminBusinessCreateSerializer(business, data=request.data, partial=True)
    if not serializer.is_valid():
        first = next(iter(serializer.errors.values()))
        message = first[0] if isinstance(first, list) else first
        return Response({'error': str(message)}, status=400)
    business = serializer.save()
    return Response({
        'id': business.id,
        'name': business.name,
        'slug': business.slug,
        'status': business.status,
        'is_featured': business.is_featured,
        'is_verified': business.is_verified,
        'accepts_orders': business.accepts_orders,
        'owner_email': business.owner.email,
    })


@api_view(['POST'])
@permission_classes([IsAdminUser])
def admin_create_business(request):
    from apps.businesses.serializers import AdminBusinessCreateSerializer
    serializer = AdminBusinessCreateSerializer(data=request.data)
    if not serializer.is_valid():
        first = next(iter(serializer.errors.values()))
        message = first[0] if isinstance(first, list) else first
        return Response({'error': str(message)}, status=400)
    business = serializer.save()
    return Response({
        'id': business.id,
        'name': business.name,
        'slug': business.slug,
        'owner_email': business.owner.email,
        'status': business.status,
    }, status=201)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def admin_business_import_template(request):
    from openpyxl import Workbook
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = 'Businesses'
    headers = [
        'name', 'category_slug', 'phone', 'city', 'address', 'state', 'description',
        'short_description', 'latitude', 'longitude', 'whatsapp', 'email', 'website',
        'facebook', 'instagram', 'payment_mode', 'card_payment', 'free_wifi', 'smoking',
        'offer_package', 'is_registered', 'has_parking', 'offer_delivery', 'owner_email',
        'logo_filename', 'banner_filename',
    ]
    ws.append(headers)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="vbazaar-bulk-import-template.xlsx"'
    wb.save(response)
    return response


@api_view(['POST'])
@permission_classes([IsAdminUser])
def admin_bulk_import_businesses(request):
    from openpyxl import load_workbook
    from django.db import transaction
    from apps.businesses.models import BusinessCategory
    from apps.businesses.serializers import _unique_slug
    from apps.accounts.services import get_or_create_dummy_business_owner

    upload = request.FILES.get('file')
    if not upload:
        return Response({'error': 'No .xlsx file uploaded (field name: "file").'}, status=400)
    try:
        wb = load_workbook(upload, data_only=True)
    except Exception:
        return Response({'error': 'Could not read the uploaded file — is it a valid .xlsx?'}, status=400)
    ws = wb.active

    images_by_name = {f.name: f for f in request.FILES.getlist('images')}
    header = [c.value for c in ws[1]]
    col = {name: idx for idx, name in enumerate(header) if name}

    def cell(row, key, default=''):
        idx = col.get(key)
        if idx is None or idx >= len(row):
            return default
        val = row[idx].value
        return val if val is not None else default

    def tri(v):
        v = str(v).strip().lower() if v not in (None, '') else ''
        if v in ('yes', 'true', '1'):
            return Business.TriState.YES
        if v in ('no', 'false', '0'):
            return Business.TriState.NO
        return Business.TriState.NA

    def nullable_bool(v):
        v = str(v).strip().lower() if v not in (None, '') else ''
        if v in ('yes', 'true', '1'):
            return True
        if v in ('no', 'false', '0'):
            return False
        return None

    rows_out = []
    created_count = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if all(c.value in (None, '') for c in row):
            continue

        errors = []
        name = str(cell(row, 'name', '')).strip()
        if not name:
            errors.append('Missing required field: name')

        category_slug = str(cell(row, 'category_slug', '')).strip()
        category = BusinessCategory.objects.filter(slug=category_slug).first() if category_slug else None
        if not category_slug:
            errors.append('Missing required field: category_slug')
        elif not category:
            errors.append(f'Category slug "{category_slug}" not found.')

        try:
            lat = float(cell(row, 'latitude'))
            lon = float(cell(row, 'longitude'))
        except (TypeError, ValueError):
            lat = lon = None

        logo_file = banner_file = None
        logo_fn = str(cell(row, 'logo_filename', '')).strip()
        if logo_fn:
            logo_file = images_by_name.get(logo_fn)
            if not logo_file:
                errors.append(f'Referenced logo file "{logo_fn}" was not found among uploaded images.')
        banner_fn = str(cell(row, 'banner_filename', '')).strip()
        if banner_fn:
            banner_file = images_by_name.get(banner_fn)
            if not banner_file:
                errors.append(f'Referenced banner file "{banner_fn}" was not found among uploaded images.')

        phone = str(cell(row, 'phone', '')).strip()
        if not phone:
            errors.append('Missing required field: phone')

        if errors:
            rows_out.append({'row': row_num, 'name': name, 'status': 'error', 'errors': errors})
            continue

        try:
            with transaction.atomic():
                owner, owner_created = get_or_create_dummy_business_owner(
                    phone=phone, email=str(cell(row, 'owner_email', '')).strip() or None,
                )
                business = Business(
                    owner=owner, name=name, slug=_unique_slug(name), category=category,
                    description=str(cell(row, 'description', '')) or 'No description provided.',
                    short_description=str(cell(row, 'short_description', '')),
                    address=str(cell(row, 'address', '')) or str(cell(row, 'city', '')) or 'Address pending',
                    city=str(cell(row, 'city', '')) or 'Pokhara',
                    state=str(cell(row, 'state', '')),
                    latitude=lat, longitude=lon, phone=phone,
                    whatsapp=str(cell(row, 'whatsapp', '')), email=str(cell(row, 'email', '')),
                    website=str(cell(row, 'website', '')), facebook=str(cell(row, 'facebook', '')),
                    instagram=str(cell(row, 'instagram', '')),
                    payment_mode=(str(cell(row, 'payment_mode', '')).strip().lower() or Business.PaymentMode.CASH_ONLY),
                    card_payment=tri(cell(row, 'card_payment')),
                    free_wifi=tri(cell(row, 'free_wifi')),
                    smoking=tri(cell(row, 'smoking')),
                    offer_package=tri(cell(row, 'offer_package')),
                    is_registered=nullable_bool(cell(row, 'is_registered')),
                    has_parking=nullable_bool(cell(row, 'has_parking')),
                    offer_delivery=nullable_bool(cell(row, 'offer_delivery')) or False,
                    status='active',
                )
                if logo_file:
                    business.logo = logo_file
                if banner_file:
                    business.banner = banner_file
                business.full_clean(exclude=['slug', 'owner'])
                business.save()
            rows_out.append({
                'row': row_num, 'name': name, 'status': 'success',
                'business_id': business.id, 'slug': business.slug,
                'owner_email': owner.email, 'owner_created': owner_created,
            })
            created_count += 1
        except Exception as e:
            rows_out.append({'row': row_num, 'name': name, 'status': 'error', 'errors': [str(e)]})

    return Response({
        'summary': {'total': len(rows_out), 'created': created_count, 'failed': len(rows_out) - created_count},
        'rows': rows_out,
    })


@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def admin_delete_business(request, pk):
    try:
        business = Business.objects.get(pk=pk)
        business.delete()
        return Response({'message': 'Business deleted.'})
    except Business.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def admin_archive_business(request, pk):
    try:
        business = Business.objects.get(pk=pk)
    except Business.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)
    business.status = Business.Status.ARCHIVED
    business.save(update_fields=['status'])
    if request.data.get('archive_branches'):
        business.branches.exclude(status=Business.Status.ARCHIVED).update(status=Business.Status.ARCHIVED)
    return Response({'id': business.id, 'status': business.status, 'branch_count': business.branches.count()})


@api_view(['POST'])
@permission_classes([IsAdminUser])
def admin_restore_business(request, pk):
    try:
        business = Business.objects.get(pk=pk)
    except Business.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)
    business.status = Business.Status.ACTIVE
    business.save(update_fields=['status'])
    return Response({'id': business.id, 'status': business.status})


@api_view(['GET'])
@permission_classes([IsAdminUser])
def admin_business_qr(request, pk):
    try:
        business = Business.objects.get(pk=pk)
    except Business.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)
    buf = generate_qr_card(business, business_target_url(business, request))
    return HttpResponse(buf.getvalue(), content_type='image/png',
                         headers={'Content-Disposition': f'attachment; filename="{business.slug}-qr.png"'})


@api_view(['GET'])
@permission_classes([IsAdminUser])
def admin_business_qr_bulk(request):
    """?ids=1,2,3 for a hand-picked subset, omitted for every business."""
    qs = Business.objects.all()
    ids_param = request.query_params.get('ids')
    if ids_param:
        ids = [int(x) for x in ids_param.split(',') if x.strip().isdigit()]
        qs = qs.filter(id__in=ids)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w') as zf:
        for b in qs:
            png = generate_qr_card(b, business_target_url(b, request))
            zf.writestr(f"{b.slug}-qr.png", png.getvalue())
    buf.seek(0)
    return HttpResponse(buf.getvalue(), content_type='application/zip',
                         headers={'Content-Disposition': 'attachment; filename="vbazaar-qr-codes.zip"'})


@api_view(['GET'])
@permission_classes([IsAdminUser])
def admin_business_export(request):
    """?status= and ?search= mirror admin_businesses' own filtering."""
    status_filter = request.query_params.get('status', '')
    search = request.query_params.get('search', '')
    qs = Business.objects.all().order_by('name')
    if status_filter:
        qs = qs.filter(status=status_filter)
    if search:
        q = Q(name__icontains=search)
        if search.isdigit():
            q |= Q(id=int(search))
        qs = qs.filter(q)
    buf = businesses_to_excel(qs)
    return HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="vbazaar-businesses.xlsx"'},
    )


@api_view(['GET'])
@permission_classes([IsAdminUser])
def admin_users(request):
    role = request.query_params.get('role', '')
    search = request.query_params.get('search', '')
    qs = User.objects.all()
    if role:
        qs = qs.filter(role=role)
    if search:
        qs = qs.filter(email__icontains=search) | qs.filter(first_name__icontains=search)
    qs = qs.order_by('-date_joined')

    data = []
    for u in qs:
        data.append({
            'id': u.id,
            'email': u.email,
            'full_name': u.get_full_name(),
            'username': u.username,
            'role': u.role,
            'is_active': u.is_active,
            'is_verified': u.is_verified,
            'phone': u.phone,
            'date_joined': u.date_joined.isoformat(),
        })
    return Response(data)


@api_view(['PATCH'])
@permission_classes([IsAdminUser])
def admin_update_user(request, pk):
    try:
        user = User.objects.get(pk=pk)
        if 'is_active' in request.data:
            user.is_active = request.data['is_active']
        if 'is_verified' in request.data:
            user.is_verified = request.data['is_verified']
        user.save()
        return Response({'id': user.id, 'is_active': user.is_active})
    except User.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def admin_reset_user_password(request, pk):
    """Super Admin resets any Business/Branch Admin's password. Returns the new
    password once — it is never stored or displayable again."""
    import secrets
    try:
        user = User.objects.get(pk=pk)
    except User.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)
    new_password = secrets.token_urlsafe(9)
    user.set_password(new_password)
    user.save(update_fields=['password'])
    return Response({'email': user.email, 'phone': user.phone, 'new_password': new_password})


@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def admin_delete_user(request, pk):
    if request.user.pk == pk:
        return Response({'error': "You can't delete your own account while logged in as it."}, status=400)
    try:
        user = User.objects.get(pk=pk)
        business_count = user.businesses.count()
        user.delete()
        return Response({'message': f'User deleted.{f" {business_count} business(es) owned by this user were also deleted." if business_count else ""}'})
    except User.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def admin_orders(request):
    status_filter = request.query_params.get('status', '')
    qs = Order.objects.all().select_related('customer', 'business').order_by('-created_at')[:100]
    if status_filter:
        qs = Order.objects.filter(status=status_filter).select_related('customer', 'business').order_by('-created_at')[:100]

    data = []
    for o in qs:
        data.append({
            'id': o.id,
            'customer_email': o.customer.email,
            'customer_name': o.customer.get_full_name(),
            'business_name': o.business.name,
            'status': o.status,
            'total': str(o.total),
            'payment_method': o.payment_method,
            'is_paid': o.is_paid,
            'created_at': o.created_at.isoformat(),
        })
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def admin_reviews(request):
    qs = Review.objects.all().select_related('customer', 'business').order_by('-created_at')[:100]
    data = []
    for r in qs:
        data.append({
            'id': r.id,
            'customer_name': r.customer.get_full_name(),
            'customer_email': r.customer.email,
            'business_name': r.business.name,
            'rating': r.rating,
            'title': r.title,
            'content': r.content,
            'is_approved': r.is_approved,
            'created_at': r.created_at.isoformat(),
        })
    return Response(data)


@api_view(['PATCH'])
@permission_classes([IsAdminUser])
def admin_update_review(request, pk):
    try:
        review = Review.objects.get(pk=pk)
        if 'is_approved' in request.data:
            review.is_approved = request.data['is_approved']
        review.save()
        return Response({'id': review.id, 'is_approved': review.is_approved})
    except Review.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)


@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def admin_delete_review(request, pk):
    try:
        Review.objects.get(pk=pk).delete()
        return Response({'message': 'Review deleted.'})
    except Review.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def admin_categories(request):
    cats = BusinessCategory.objects.all()
    data = [{'id': c.id, 'name': c.name, 'slug': c.slug, 'icon': c.icon, 'order': c.order,
             'category_code': c.category_code, 'business_count': c.businesses.count()} for c in cats]
    return Response(data)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def admin_create_category(request):
    from django.utils.text import slugify
    name = request.data.get('name', '')
    icon = request.data.get('icon', '🏪')
    order = request.data.get('order', 0)
    if not name:
        return Response({'error': 'Name is required.'}, status=400)
    slug = slugify(name)
    cat = BusinessCategory.objects.create(name=name, slug=slug, icon=icon, order=order)
    return Response({'id': cat.id, 'name': cat.name, 'slug': cat.slug, 'icon': cat.icon})


@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def admin_delete_category(request, pk):
    try:
        BusinessCategory.objects.get(pk=pk).delete()
        return Response({'message': 'Category deleted.'})
    except BusinessCategory.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def admin_advertisements(request):
    ads = Advertisement.objects.select_related('business').all()
    data = [{
        'id': a.id,
        'image': a.business.banner.url if a.business.banner else None,
        'business_id': a.business_id,
        'business_name': a.business.name,
        'is_active': a.is_active,
        'order': a.order,
    } for a in ads]
    return Response(data)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def admin_create_advertisement(request):
    business_id = request.data.get('business_id')
    if not business_id:
        return Response({'error': 'business_id is required.'}, status=400)
    try:
        business = Business.objects.get(pk=business_id)
    except Business.DoesNotExist:
        return Response({'error': 'Business not found.'}, status=404)
    if not business.banner:
        return Response({'error': 'This business has no cover photo uploaded yet — ask them to add one in their profile before featuring them as an ad.'}, status=400)
    ad = Advertisement.objects.create(business=business)
    return Response({'id': ad.id, 'business_name': business.name}, status=201)


@api_view(['PATCH'])
@permission_classes([IsAdminUser])
def admin_update_advertisement(request, pk):
    try:
        ad = Advertisement.objects.get(pk=pk)
        if 'is_active' in request.data:
            ad.is_active = request.data['is_active']
        if 'order' in request.data:
            ad.order = request.data['order']
        ad.save()
        return Response({'id': ad.id, 'is_active': ad.is_active, 'order': ad.order})
    except Advertisement.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)


@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def admin_delete_advertisement(request, pk):
    try:
        Advertisement.objects.get(pk=pk).delete()
        return Response({'message': 'Advertisement deleted.'})
    except Advertisement.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)